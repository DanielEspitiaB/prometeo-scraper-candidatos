"""
Evaluador determinista — Prometeo (zona determinista).

Flujo:
  1. Lee perfiles crudos de Apify desde salida_validacion.json.
  2. Arma el "perfil limpio" (solo los campos que el modelo necesita).
  3. Evalua cada perfil con Claude contra un scorecard (plantilla de prompt FIJA).
  4. El codigo (no el modelo) suma puntos y calcula la banda con scorecard.py.
  5. Escribe un Excel con formato: resultados ordenados por puntaje.

Reglas que respeta:
  - ANTHROPIC_API_KEY se LEE de .streamlit/secrets.toml, nunca va en el codigo.
  - La llamada a Claude usa un cuerpo FIJO; el codigo inyecta los datos (no un agente).
  - Claude SOLO asigna puntos + justificacion por criterio. La suma total y la
    banda las calcula el codigo (determinista), incluida la regla de eliminatorios.

Como correr (desde la carpeta "Scrapper Candidatos"):
    python evaluar.py
"""

import csv
import datetime
import io
import json
import re
import sys
import tomllib
import urllib.request
import urllib.error
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from dataclasses import asdict

import scorecard as sc

CARPETA = Path(__file__).parent
SECRETS = CARPETA / ".streamlit" / "secrets.toml"
ENTRADA = CARPETA / "salida_validacion.json"
SALIDA = CARPETA / "resultados_evaluacion.xlsx"

MODELO = "claude-sonnet-4-6"  # modelo por defecto: prioriza calidad de evaluacion (no perder buenos candidatos)
ENDPOINT_CLAUDE = "https://api.anthropic.com/v1/messages"

# Precios aproximados en USD por millon de tokens (entrada / salida)
PRECIOS_USD = {
    "claude-sonnet-4-6": (3.0, 15.0),
    "claude-haiku-4-5-20251001": (1.0, 5.0),
}


def costo_usd(modelo: str, usage: dict) -> float:
    pin, pout = PRECIOS_USD.get(modelo, (3.0, 15.0))
    return (usage.get("input_tokens", 0) * pin + usage.get("output_tokens", 0) * pout) / 1_000_000

# Enricher ACTIVO: harvestapi (más barato $0.004, sin email, datos más completos)
ENRICHER_ACTOR = "harvestapi~linkedin-profile-scraper"
ENRICHER_ENDPOINT = f"https://api.apify.com/v2/acts/{ENRICHER_ACTOR}/run-sync-get-dataset-items"
ENRICHER_MODE = "Profile details no email ($4 per 1k)"

# dev_fusion queda disponible (para email de finalistas si se necesita)
APIFY_ACTOR = "dev_fusion~linkedin-profile-scraper"
APIFY_ENDPOINT = f"https://api.apify.com/v2/acts/{APIFY_ACTOR}/run-sync-get-dataset-items"

# Colores por banda para el Excel (formato ARGB de openpyxl)
COLORES_BANDA = {
    "Ideal": "C6EFCE",          # verde
    "Avanzar": "D9EAD3",        # verde claro
    "Pendiente": "FFEB9C",      # amarillo
    "No recomendado": "FFC7CE", # rojo
}


# ---------------------------------------------------------------------------
# Secrets
# ---------------------------------------------------------------------------
def leer_api_key() -> str:
    if not SECRETS.exists():
        sys.exit(f"No encuentro el archivo de secrets en: {SECRETS}")
    data = tomllib.load(open(SECRETS, "rb"))
    key = data.get("ANTHROPIC_API_KEY", "")
    if not key or key.startswith("PEGA_AQUI"):
        sys.exit("Falta tu ANTHROPIC_API_KEY en .streamlit/secrets.toml.")
    return key


# ---------------------------------------------------------------------------
# Perfil limpio: de ~70 campos crudos a solo lo que el modelo necesita
# ---------------------------------------------------------------------------
def _ahora_frac() -> float:
    hoy = datetime.date.today()
    return hoy.year + (hoy.month - 1) / 12


def perfil_limpio(p: dict) -> dict:
    """Normaliza un perfil crudo (dev_fusion o harvestapi) al mismo formato limpio."""
    if "experiences" in p:  # dev_fusion usa 'experiences' (plural)
        return _limpio_devfusion(p)
    return _limpio_harvest(p)  # harvestapi usa 'experience' (singular)


_MESES = {"Jan": 1, "Feb": 2, "Mar": 3, "Apr": 4, "May": 5, "Jun": 6,
          "Jul": 7, "Aug": 8, "Sep": 9, "Oct": 10, "Nov": 11, "Dec": 12}


def _mes_idx(d):
    """Convierte {month, year} a un índice de meses. None si no hay año."""
    if not isinstance(d, dict) or not d.get("year"):
        return None
    return d["year"] * 12 + (_MESES.get(d.get("month"), 1) - 1)


def _anios_harvest(exp: list):
    """Años de experiencia = UNIÓN de los periodos trabajados (sin contar solapes)."""
    hoy = datetime.date.today()
    ahora = hoy.year * 12 + (hoy.month - 1)
    intervalos = []
    for e in exp:
        ini = _mes_idx(e.get("startDate"))
        if ini is None:
            continue
        fin = _mes_idx(e.get("endDate"))
        if fin is None:  # "Present" o sin fecha de fin
            fin = ahora
        intervalos.append((ini, max(fin, ini)))
    if not intervalos:
        return None
    intervalos.sort()
    total, cs, ce = 0, *intervalos[0]
    for s, e in intervalos[1:]:
        if s <= ce:
            ce = max(ce, e)
        else:
            total += ce - cs
            cs, ce = s, e
    total += ce - cs
    return round(total / 12, 1)


def _limpio_harvest(p: dict) -> dict:
    exp = p.get("experience") or []
    nombre = (str(p.get("firstName") or "") + " " + str(p.get("lastName") or "")).strip() or (p.get("headline") or "")
    loc = p.get("location") or {}
    ubic = (loc.get("parsed") or {}).get("text") or loc.get("linkedinText") or ""
    anios = _anios_harvest(exp)
    return {
        "nombre": nombre,
        "titular": p.get("headline"),
        "resumen": (p.get("about") or "")[:600],
        "ubicacion": ubic,
        "anios_experiencia_total": anios,
        "cargo_actual": exp[0].get("position") if exp else None,
        "experiencias": [
            {
                "cargo": e.get("position"),
                "empresa": e.get("companyName"),
                "industria": e.get("companyIndustry"),
                "desde": (e.get("startDate") or {}).get("text"),
                "hasta": (e.get("endDate") or {}).get("text") or "actual",
                "sigue_ahi": ((e.get("endDate") or {}).get("text") or "").lower() in ("", "present"),
                "descripcion": (e.get("description") or "")[:400],
            }
            for e in exp
        ],
        "educacion": [
            {"institucion": ed.get("schoolName"), "titulo": ed.get("degree") or ed.get("fieldOfStudy")}
            for ed in (p.get("education") or [])
        ],
        "skills": [s.get("name") if isinstance(s, dict) else s for s in (p.get("skills") or [])],
    }


def _limpio_devfusion(p: dict) -> dict:
    return {
        "nombre": p.get("fullName"),
        "titular": p.get("headline"),
        "resumen": (p.get("about") or "")[:600],
        "ubicacion": p.get("addressWithCountry"),
        "anios_experiencia_total": p.get("totalExperienceYears"),
        "cargo_actual": p.get("jobTitle"),
        "experiencias": [
            {
                "cargo": e.get("title"),
                "empresa": e.get("companyName"),
                "industria": e.get("companyIndustry"),
                "desde": e.get("jobStartedOn"),
                "hasta": e.get("jobEndedOn"),
                "sigue_ahi": e.get("jobStillWorking"),
                "descripcion": (e.get("jobDescription") or "")[:400],
            }
            for e in (p.get("experiences") or [])
        ],
        "educacion": [
            {"institucion": e.get("title"), "titulo": e.get("subtitle")}
            for e in (p.get("educations") or [])
        ],
        "skills": [s.get("title") if isinstance(s, dict) else s for s in (p.get("skills") or [])],
    }


def datos_contacto(p: dict) -> dict:
    """Datos que van al Excel para el reclutador (NO al modelo)."""
    if "experiences" in p:  # dev_fusion
        return {
            "email": p.get("email") or "",
            "telefono": p.get("mobileNumber") or "",
            "url": p.get("linkedinUrl") or p.get("linkedinPublicUrl") or "",
        }
    # harvestapi (emails en lista; en modo solo-detalles viene vacío)
    emails = p.get("emails") or []
    correo = emails[0] if emails else ""
    if isinstance(correo, dict):
        correo = correo.get("email") or correo.get("value") or ""
    return {"email": correo, "telefono": "", "url": p.get("linkedinUrl") or ""}


# ---------------------------------------------------------------------------
# Raspado de perfiles por URL (Apify) — cuerpo de peticion FIJO
# ---------------------------------------------------------------------------
def _raspar_lote(lote: list, apify_token: str) -> list:
    """Una llamada al enricher harvestapi (solo detalles, sin email). Body fijo."""
    body = json.dumps({"profileScraperMode": ENRICHER_MODE, "urls": lote}).encode("utf-8")
    req = urllib.request.Request(
        f"{ENRICHER_ENDPOINT}?token={apify_token}", data=body, method="POST",
        headers={"Content-Type": "application/json"},
    )
    with urllib.request.urlopen(req, timeout=300) as resp:
        return json.loads(resp.read().decode("utf-8"))


def scrape_perfiles(urls: list, apify_token: str, chunk: int = 100,
                    max_workers: int = 8, on_progress=None) -> list:
    """Enriquece URLs de LinkedIn con dev_fusion, en lotes EN PARALELO (rapido y sin timeouts).

    Divide en lotes de 'chunk' y los lanza simultaneamente (hasta 'max_workers' a la vez).
    on_progress(hechos, total) actualiza el avance a medida que cada lote termina.
    """
    urls = list(urls)
    lotes = [urls[i:i + chunk] for i in range(0, len(urls), chunk)]
    resultados, errores, hechos = [], [], 0

    with ThreadPoolExecutor(max_workers=min(max_workers, len(lotes) or 1)) as ex:
        futuros = {ex.submit(_raspar_lote, lote, apify_token): lote for lote in lotes}
        for fut in as_completed(futuros):
            lote = futuros[fut]
            hechos += len(lote)
            try:
                items = fut.result()
                resultados.extend(it for it in items if "error" not in it)
                errores.extend(it["error"] for it in items if "error" in it)
            except Exception as e:
                errores.append(str(e))
            if on_progress:
                on_progress(min(hechos, len(urls)), len(urls))

    if not resultados and errores:
        raise RuntimeError(errores[0])
    return resultados


# ---------------------------------------------------------------------------
# Prompt FIJO para Claude
# ---------------------------------------------------------------------------
def construir_prompt(scorecard: sc.Scorecard, perfil: dict) -> str:
    criterios = [
        {"id": c.id, "criterio": c.label, "tipo": c.type,
         "puntos_maximos": c.points, "nota": c.notes}
        for c in scorecard.criteria
    ]
    return (
        "Eres un evaluador experto de talento para una consultora de headhunting. "
        "Recibes un SCORECARD con criterios y un PERFIL de LinkedIn ya resumido.\n\n"
        "Tu tarea: para CADA criterio, decide cuantos puntos merece el candidato "
        "(entre 0 y 'puntos_maximos'), basandote en la evidencia del perfil "
        "(titular, resumen, experiencias, skills, educacion). Sé estricto y justo.\n\n"
        "Reglas:\n"
        "- Para criterios de tipo 'eliminatory': si el candidato NO cumple el requisito, "
        "pon \"cumple\": false (eso lo descalifica). Si cumple, \"cumple\": true.\n"
        "- Para tipo 'scored': \"cumple\" es informativo; lo importante son los puntos.\n"
        "- Si no hay evidencia suficiente de un criterio, otorga pocos o cero puntos.\n"
        "- La justificacion debe ser UNA frase corta citando la evidencia concreta.\n\n"
        "Responde UNICAMENTE con JSON valido, sin texto extra ni markdown, con esta forma exacta:\n"
        '{"criterios": [{"id": "...", "puntos": 0, "cumple": true, "justificacion": "..."}]}\n\n'
        f"SCORECARD (rol: {scorecard.role}, ubicacion buscada: {scorecard.location}):\n"
        f"{json.dumps(criterios, ensure_ascii=False, indent=2)}\n\n"
        f"PERFIL DEL CANDIDATO:\n{json.dumps(perfil, ensure_ascii=False, indent=2)}"
    )


def llamar_claude(prompt: str, api_key: str, modelo: str = MODELO):
    body = json.dumps({
        "model": modelo,
        "max_tokens": 1500,
        "messages": [{"role": "user", "content": prompt}],
    }).encode("utf-8")
    req = urllib.request.Request(
        ENDPOINT_CLAUDE, data=body, method="POST",
        headers={
            "x-api-key": api_key,
            "anthropic-version": "2023-06-01",
            "content-type": "application/json",
        },
    )
    try:
        with urllib.request.urlopen(req, timeout=120) as resp:
            data = json.loads(resp.read().decode("utf-8"))
    except urllib.error.HTTPError as e:
        sys.exit(f"Claude respondio error HTTP {e.code}:\n{e.read().decode('utf-8', 'replace')}")
    texto = data["content"][0]["text"]
    usage = data.get("usage", {})
    # Extraer el bloque JSON de forma robusta
    m = re.search(r"\{.*\}", texto, re.DOTALL)
    if not m:
        raise ValueError(f"Claude no devolvio JSON:\n{texto}")
    return json.loads(m.group(0)), usage


# ---------------------------------------------------------------------------
# Evaluacion + banda (el codigo hace las cuentas, no el modelo)
# ---------------------------------------------------------------------------
def evaluar_perfil(scorecard: sc.Scorecard, crudo: dict, api_key: str, modelo: str = MODELO) -> dict:
    limpio = perfil_limpio(crudo)
    respuesta, usage = llamar_claude(construir_prompt(scorecard, limpio), api_key, modelo)

    por_id = {c.id: c for c in scorecard.criteria}
    puntos = {}
    justificaciones = []
    descalificado = False

    for item in respuesta.get("criterios", []):
        cid = item.get("id")
        crit = por_id.get(cid)
        if not crit:
            continue
        p = max(0, min(int(item.get("puntos", 0)), crit.points))  # nunca supera el maximo
        puntos[cid] = p
        if crit.type == "eliminatory" and item.get("cumple") is False:
            descalificado = True
        justificaciones.append(f"[{crit.label}] {item.get('justificacion', '')}")

    total = sum(puntos.values())
    banda = scorecard.band_for(total, disqualified=descalificado)
    return {
        "nombre": limpio["nombre"],
        "puntos": puntos,
        "total": total,
        "descalificado": descalificado,
        "banda": banda,
        "justificacion": "  •  ".join(justificaciones),
        "contacto": datos_contacto(crudo),
        "usage": usage,
        "modelo": modelo,
    }


# ---------------------------------------------------------------------------
# Excel con formato
# ---------------------------------------------------------------------------
def construir_workbook(scorecard: sc.Scorecard, resultados: list):
    """Construye el libro de Excel en memoria (lo usa tanto el script como la app)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Resultados"

    criterios = scorecard.criteria
    encabezados = ["Candidato", "Banda", "Puntaje total"] + \
        [c.label for c in criterios] + ["Justificación", "Email", "Teléfono", "LinkedIn"]
    ws.append(encabezados)
    for celda in ws[1]:
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill("solid", fgColor="404040")
        celda.alignment = Alignment(vertical="center", wrap_text=True)

    # Ordenar de mejor a peor puntaje
    for r in sorted(resultados, key=lambda x: x["total"], reverse=True):
        fila = [r["nombre"], r["banda"], r["total"]] + \
            [r["puntos"].get(c.id, 0) for c in criterios] + \
            [r["justificacion"], r["contacto"]["email"], r["contacto"]["telefono"], r["contacto"]["url"]]
        ws.append(fila)
        celda_banda = ws.cell(row=ws.max_row, column=2)
        color = COLORES_BANDA.get(r["banda"])
        if color:
            celda_banda.fill = PatternFill("solid", fgColor=color)
            celda_banda.font = Font(bold=True)

    # Anchos de columna basicos
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 12
    return wb


def escribir_excel(scorecard: sc.Scorecard, resultados: list):
    """Guarda el Excel en disco (para el script standalone)."""
    construir_workbook(scorecard, resultados).save(SALIDA)


def _experiencia_str(e: dict) -> str:
    """Una experiencia: 'Cargo @ Empresa (fechas): descripcion de tareas' (si existe)."""
    base = f"{e['cargo']} @ {e['empresa'] or '?'} ({e['desde'] or '?'}–{e['hasta'] or 'actual'})"
    desc = (e.get("descripcion") or "").strip()
    return f"{base}: {desc}" if desc else base


# Caracteres de control ilegales que Excel/xlsx no acepta (los quitamos antes de escribir)
_CARACTERES_ILEGALES = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")


def _limpiar_celda(v):
    """Quita caracteres de control invisibles que rompen Excel. Deja números/None igual."""
    return _CARACTERES_ILEGALES.sub("", v) if isinstance(v, str) else v


def filas_datos(crudos: list) -> list:
    """Filas (dicts) con TODOS los datos de cada candidato. Base comun para Excel y CSV."""
    filas = []
    for crudo in crudos:
        p = perfil_limpio(crudo)
        c = datos_contacto(crudo)
        exps = p["experiencias"]
        empresa_actual = next((e.get("empresa") for e in exps if e.get("sigue_ahi") and e.get("empresa")), "")
        if not empresa_actual and exps:
            empresa_actual = exps[0].get("empresa") or ""
        fila = {
            "Nombre": p["nombre"],
            "Titular": p["titular"],
            "Ubicación": p["ubicacion"],
            "Años exp.": p["anios_experiencia_total"],
            "Cargo actual": p["cargo_actual"],
            "Empresa actual": empresa_actual,
            "Email": c["email"],
            "Teléfono": c["telefono"],
            "LinkedIn": c["url"],
            "Skills": ", ".join(str(s) for s in p["skills"]),
            "Educación": " | ".join(
                f"{e['institucion']} ({e['titulo']})" if e.get("titulo") else (e.get("institucion") or "")
                for e in p["educacion"] if e.get("institucion")
            ),
            "Experiencia": "\n".join(_experiencia_str(e) for e in exps if e.get("cargo")),
        }
        filas.append({k: _limpiar_celda(v) for k, v in fila.items()})
    return filas


def construir_workbook_datos(crudos: list):
    """Excel con los DATOS de cada candidato (sin evaluacion). Branding Prometeo (navy)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    filas = filas_datos(crudos)
    wb = Workbook()
    ws = wb.active
    ws.title = "Candidatos"
    encabezados = list(filas[0].keys()) if filas else ["Nombre"]
    ws.append(encabezados)
    for celda in ws[1]:
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill("solid", fgColor="142649")  # azul marino Prometeo
        celda.alignment = Alignment(vertical="center", wrap_text=True)
    for f in filas:
        ws.append(list(f.values()))

    # Ajuste de linea en las columnas largas (Skills, Educación, Experiencia)
    wrap = Alignment(wrap_text=True, vertical="top")
    for fila in ws.iter_rows(min_row=2, min_col=10, max_col=12):
        for celda in fila:
            celda.alignment = wrap

    anchos = {"A": 22, "B": 42, "C": 22, "D": 9, "E": 24, "F": 24,
              "G": 26, "H": 16, "I": 38, "J": 40, "K": 40, "L": 70}
    for col, w in anchos.items():
        ws.column_dimensions[col].width = w
    return wb


def construir_csv_datos(crudos: list) -> bytes:
    """CSV con TODOS los datos de cada candidato (mismo contenido que el Excel)."""
    filas = filas_datos(crudos)
    campos = list(filas[0].keys()) if filas else ["Nombre"]
    buf = io.StringIO()
    w = csv.DictWriter(buf, fieldnames=campos)
    w.writeheader()
    for f in filas:
        w.writerow(f)
    # utf-8-sig para que Excel abra bien los acentos
    return buf.getvalue().encode("utf-8-sig")


# ---------------------------------------------------------------------------
# Principal
# ---------------------------------------------------------------------------
def main():
    api_key = leer_api_key()
    if not ENTRADA.exists():
        sys.exit(f"No encuentro {ENTRADA.name}. Corre primero validar_apify.py.")
    crudos = json.load(open(ENTRADA, encoding="utf-8"))
    crudos = [c for c in crudos if "error" not in c]  # ignora respuestas de error

    scorecard = sc.senior_go_developer()  # scorecard de prueba
    print(f"Evaluando {len(crudos)} perfil(es) contra: {scorecard.role}\n")

    resultados = []
    for crudo in crudos:
        print(f"  Evaluando: {crudo.get('fullName')} ...")
        resultados.append(evaluar_perfil(scorecard, crudo, api_key))

    print("\n" + "=" * 60)
    print("RESULTADOS (de mayor a menor puntaje):")
    print("=" * 60)
    for r in sorted(resultados, key=lambda x: x["total"], reverse=True):
        marca = "  [DESCALIFICADO]" if r["descalificado"] else ""
        print(f"  {r['nombre']:<24} {r['total']:>3} pts  -> {r['banda']}{marca}")

    escribir_excel(scorecard, resultados)
    print("\nExcel generado:", SALIDA.name)


if __name__ == "__main__":
    main()
